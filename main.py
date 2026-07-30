from __future__ import annotations

import argparse
import hashlib
import html as html_lib
import json
import re
import shutil
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs, urlencode, urldefrag, urljoin, urlparse, urlunparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from dateutil import parser as date_parser
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from playwright.sync_api import Browser, Page, sync_playwright
except ImportError:  # The HTTP-only path still works without Playwright.
    Browser = Any  # type: ignore[assignment,misc]
    Page = Any  # type: ignore[assignment,misc]
    sync_playwright = None


ROOT = Path(__file__).resolve().parent
CONFIG_FILE = ROOT / "config" / "companies.json"
DATA_DIR = ROOT / "data"
DASHBOARD_DATA_DIR = ROOT / "dashboard" / "data"
OUTPUT_DIR = ROOT / "output"

CURRENT_JSON = DATA_DIR / "current_jobs.json"
CURRENT_CSV = DATA_DIR / "current_jobs.csv"
STATE_JSON = DATA_DIR / "job_state.json"
HEALTH_JSON = DATA_DIR / "collector_health.json"
EXCEL_FILE = OUTPUT_DIR / "data-ai-jobs.xlsx"

DASHBOARD_JOBS_JSON = DASHBOARD_DATA_DIR / "jobs.json"
DASHBOARD_META_JSON = DASHBOARD_DATA_DIR / "meta.json"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/131.0 Safari/537.36 DataAIJobTracker/1.0"
)

TARGET_TITLE_PATTERNS = [
    r"\banalyst\b",
    r"\bdatabase administrator\b",
    r"\bdata product manager\b",
    r"\bai product manager\b",
    r"\bprompt engineer\b",
    r"\bai researcher\b",
    r"\bmachine learning scientist\b",
    r"\bdata operations\b",
    r"\bstatistician\b",
    r"\bdata analyst\b",
    r"\bsenior data analyst\b",
    r"\bdata engineer\b",
    r"\banalytics engineer\b",
    r"\bdata scientist\b",
    r"\bdata architect\b",
    r"\bdata model(?:er|ing)\b",
    r"\bdata warehouse\b",
    r"\bdata platform\b",
    r"\bdata governance\b",
    r"\bdata quality\b",
    r"\bdata visualization\b",
    r"\bbusiness intelligence\b",
    r"\bbi analyst\b",
    r"\bbi developer\b",
    r"\bpower bi\b",
    r"\btableau (?:developer|analyst)\b",
    r"\betl (?:developer|engineer)\b",
    r"\bsnowflake (?:developer|engineer|architect)\b",
    r"\bdatabricks (?:developer|engineer|architect)\b",
    r"\bpyspark (?:developer|engineer)\b",
    r"\bmachine learning\b",
    r"\bml engineer\b",
    r"\bmlops\b",
    r"\bartificial intelligence\b",
    r"\bai engineer\b",
    r"\bai architect\b",
    r"\bai developer\b",
    r"\bgenerative ai\b",
    r"\bgenai\b",
    r"\bllm\b",
    r"\bnatural language processing\b",
    r"\bnlp engineer\b",
    r"\bcomputer vision\b",
    r"\bapplied scientist\b",
    r"\bdecision scientist\b",
    r"\bquantitative analyst\b",
    r"\bproduct analyst\b",
    r"\bmarketing analyst\b",
    r"\bfinancial data analyst\b",
]
TARGET_TITLE_RE = re.compile("|".join(TARGET_TITLE_PATTERNS), re.IGNORECASE)

STRONG_AI_DESCRIPTION_RE = re.compile(
    r"\b(?:large language model|llm|generative ai|genai|machine learning|"
    r"artificial intelligence|rag|retrieval augmented generation|"
    r"computer vision|natural language processing|mlops)\b",
    re.IGNORECASE,
)
GENERIC_ENGINEERING_TITLE_RE = re.compile(
    r"\b(?:software|backend|platform|cloud|solutions?)\s+(?:engineer|architect|developer)\b",
    re.IGNORECASE,
)
EXCLUDED_TITLE_RE = re.compile(
    r"\b(?:data entry|data center|datacenter|cable technician|field technician|"
    r"help desk|desktop support|network technician|recruiter|sales representative)\b",
    re.IGNORECASE,
)

JOB_URL_HINT_RE = re.compile(
    r"(?:/job/|/jobs/|job[-_]?detail|jobdescription|job-description|"
    r"/position/|/positions/|/opening/|/openings/|/vacancy/|/vacancies/|"
    r"requisition|careersection/.+job|find_a_job/.+job-|[?&](?:job|jobid|job_id)=)",
    re.IGNORECASE,
)
LISTING_URL_RE = re.compile(
    r"(?:search-results|job-search|search-jobs|find-jobs|find-work|all-jobs|"
    r"current-openings|current-opportunities|/careers/?$|/jobs/?$)",
    re.IGNORECASE,
)
CLOSED_JOB_RE = re.compile(
    r"(?:job (?:is )?no longer available|position has been filled|job has expired|"
    r"this job is closed|posting has expired|no longer accepting applications)",
    re.IGNORECASE,
)
REMOTE_RE = re.compile(r"\b(?:remote|work from home|telecommute|virtual)\b", re.IGNORECASE)
HYBRID_RE = re.compile(r"\bhybrid\b", re.IGNORECASE)
ONSITE_RE = re.compile(r"\b(?:on[- ]site|in[- ]office|onsite)\b", re.IGNORECASE)

MAX_CANDIDATES_PER_COMPANY = 300
MAX_SITEMAP_URLS = 300
MAX_LISTING_PAGES = 10
REQUEST_TIMEOUT = 30


@dataclass
class Job:
    company: str
    job_title: str
    location: str
    work_mode: str
    employment_type: str
    date_posted: str
    status: str
    source: str
    application_url: str
    description: str
    first_seen: str = ""
    last_seen: str = ""
    is_new: bool = False
    job_id: str = ""

    def finalize(self) -> "Job":
        self.company = clean_text(self.company)
        self.job_title = clean_text(self.job_title)
        self.location = clean_text(self.location) or "See posting"
        self.work_mode = clean_text(self.work_mode) or "See posting"
        self.employment_type = clean_text(self.employment_type) or "See posting"
        self.date_posted = normalize_date(self.date_posted)
        self.status = clean_text(self.status) or "Active"
        self.source = clean_text(self.source)
        self.application_url = canonicalize_url(self.application_url)
        self.description = clean_text(self.description)[:3000]
        self.job_id = make_job_id(self.company, self.application_url, self.job_title)
        return self


@dataclass
class CompanyHealth:
    company: str
    status: str
    jobs_found: int
    method: str
    duration_seconds: float
    message: str = ""
    careers_url: str = ""


class BrowserCollector:
    """One Chromium instance shared by all companies that need rendered pages."""

    def __init__(self, enabled: bool = True) -> None:
        self.enabled = enabled and sync_playwright is not None
        self._playwright = None
        self._browser: Browser | None = None

    def __enter__(self) -> "BrowserCollector":
        if self.enabled:
            self._playwright = sync_playwright().start()
            self._browser = self._playwright.chromium.launch(
                headless=True,
                args=["--disable-dev-shm-usage", "--no-sandbox"],
            )
        return self

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        if self._browser is not None:
            self._browser.close()
        if self._playwright is not None:
            self._playwright.stop()

    def collect_links(
        self,
        url: str,
        search_term: str | None = None,
        timeout_ms: int = 45_000,
    ) -> tuple[list[tuple[str, str]], str]:
        if not self.enabled or self._browser is None:
            return [], ""

        page = self._browser.new_page(
            user_agent=USER_AGENT,
            viewport={"width": 1440, "height": 1000},
        )
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            self._accept_cookies(page)
            page.wait_for_timeout(1800)

            if search_term:
                self._try_search(page, search_term)

            self._scroll(page)
            links: list[tuple[str, str]] = []
            for frame in page.frames:
                try:
                    frame_links = frame.locator("a[href]").evaluate_all(
                        """
                        (els) => els.map((a) => ({
                          href: a.href || a.getAttribute('href') || '',
                          text: (a.innerText || a.getAttribute('aria-label') ||
                                 a.getAttribute('title') || '').trim()
                        }))
                        """
                    )
                except Exception:
                    continue
                for item in frame_links:
                    href = str(item.get("href", "")).strip()
                    text = str(item.get("text", "")).strip()
                    if href:
                        links.append((href, text))

            return dedupe_link_pairs(links), page.content()
        finally:
            page.close()

    def render_html(self, url: str, timeout_ms: int = 45_000) -> str:
        if not self.enabled or self._browser is None:
            return ""
        page = self._browser.new_page(
            user_agent=USER_AGENT,
            viewport={"width": 1440, "height": 1000},
        )
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            self._accept_cookies(page)
            page.wait_for_timeout(1200)
            self._scroll(page, rounds=2)
            return page.content()
        finally:
            page.close()

    @staticmethod
    def _accept_cookies(page: Page) -> None:
        for label in [
            "Accept all",
            "Accept All",
            "Accept cookies",
            "Allow all",
            "I agree",
        ]:
            try:
                locator = page.get_by_role("button", name=re.compile(label, re.I))
                if locator.count() and locator.first.is_visible():
                    locator.first.click(timeout=1200)
                    break
            except Exception:
                continue

    @staticmethod
    def _try_search(page: Page, search_term: str) -> None:
        selectors = [
            "input[type='search']",
            "input[placeholder*='search' i]",
            "input[placeholder*='keyword' i]",
            "input[name*='keyword' i]",
            "input[id*='keyword' i]",
            "input[name*='search' i]",
            "input[id*='search' i]",
            "input[data-automation-id*='search' i]",
        ]
        for selector in selectors:
            try:
                locator = page.locator(selector)
                if not locator.count():
                    continue
                target = None
                for index in range(min(locator.count(), 8)):
                    candidate = locator.nth(index)
                    if candidate.is_visible():
                        target = candidate
                        break
                if target is None:
                    continue
                target.fill(search_term, timeout=1800)
                target.press("Enter")
                page.wait_for_timeout(2800)
                return
            except Exception:
                continue

    @staticmethod
    def _scroll(page: Page, rounds: int = 5) -> None:
        for _ in range(rounds):
            try:
                page.mouse.wheel(0, 5000)
                page.wait_for_timeout(600)
            except Exception:
                break


def create_session() -> requests.Session:
    retry = Retry(
        total=2,
        connect=2,
        read=2,
        backoff_factor=0.8,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET", "HEAD"}),
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
    session = requests.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/json;q=0.9,*/*;q=0.8",
        }
    )
    return session


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    raw = str(value)
    if "<" in raw and ">" in raw:
        text = BeautifulSoup(raw, "html.parser").get_text(" ", strip=True)
    else:
        text = raw
    text = html_lib.unescape(text)
    return re.sub(r"\s+", " ", text).strip(" \t\r\n-|•")


def normalize_date(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    try:
        parsed = date_parser.parse(text)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc).date().isoformat()
    except (ValueError, TypeError, OverflowError):
        return text[:40]


def canonicalize_url(url: str) -> str:
    if not url:
        return ""
    url, _ = urldefrag(url.strip())
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return ""
    query = parse_qs(parsed.query, keep_blank_values=True)
    tracking_keys = {
        "utm_source",
        "utm_medium",
        "utm_campaign",
        "utm_term",
        "utm_content",
        "source",
        "src",
        "ref",
        "referrer",
    }
    kept = {k: v for k, v in query.items() if k.lower() not in tracking_keys}
    clean_query = urlencode(kept, doseq=True)
    path = re.sub(r"/{2,}", "/", parsed.path or "/")
    return urlunparse((parsed.scheme.lower(), parsed.netloc.lower(), path.rstrip("/") or "/", "", clean_query, ""))


def make_job_id(company: str, url: str, title: str) -> str:
    raw = f"{company.casefold()}|{canonicalize_url(url)}|{title.casefold()}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:20]


def is_target_job(title: str, description: str = "") -> bool:
    title = clean_text(title)
    description = clean_text(description)
    if not title or EXCLUDED_TITLE_RE.search(title):
        return False
    if TARGET_TITLE_RE.search(title):
        return True
    return bool(
        GENERIC_ENGINEERING_TITLE_RE.search(title)
        and STRONG_AI_DESCRIPTION_RE.search(description[:2500])
    )


def detect_work_mode(*values: Any) -> str:
    text = " ".join(clean_text(value) for value in values)
    if HYBRID_RE.search(text):
        return "Hybrid"
    if REMOTE_RE.search(text):
        return "Remote"
    if ONSITE_RE.search(text):
        return "On-site"
    return "See posting"


def parse_location(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, list):
        locations = [parse_location(item) for item in value]
        return "; ".join(item for item in locations if item)
    if isinstance(value, str):
        return clean_text(value)
    if not isinstance(value, dict):
        return clean_text(value)

    address = value.get("address", value)
    if isinstance(address, list):
        return parse_location(address)
    if not isinstance(address, dict):
        return clean_text(address)

    parts = [
        address.get("addressLocality"),
        address.get("addressRegion"),
        address.get("postalCode"),
        address.get("addressCountry"),
    ]
    return ", ".join(clean_text(part) for part in parts if clean_text(part))


def is_expired(valid_through: Any) -> bool:
    text = clean_text(valid_through)
    if not text:
        return False
    try:
        parsed = date_parser.parse(text)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc) < datetime.now(timezone.utc)
    except (ValueError, TypeError, OverflowError):
        return False


def iter_json_objects(value: Any) -> Iterable[dict[str, Any]]:
    if isinstance(value, dict):
        yield value
        graph = value.get("@graph")
        if graph:
            yield from iter_json_objects(graph)
    elif isinstance(value, list):
        for item in value:
            yield from iter_json_objects(item)


def object_is_job_posting(obj: dict[str, Any]) -> bool:
    value = obj.get("@type", "")
    if isinstance(value, list):
        return any(str(item).casefold() == "jobposting" for item in value)
    return str(value).casefold() == "jobposting"


def extract_jsonld_jobs(
    html: str,
    page_url: str,
    company_name: str,
    source: str,
) -> list[Job]:
    soup = BeautifulSoup(html, "html.parser")
    jobs: list[Job] = []

    for script in soup.select("script[type='application/ld+json']"):
        raw = script.string or script.get_text("", strip=True)
        if not raw:
            continue
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            raw = raw.strip().rstrip(";")
            try:
                parsed = json.loads(raw)
            except json.JSONDecodeError:
                continue

        for obj in iter_json_objects(parsed):
            if not object_is_job_posting(obj):
                continue
            if is_expired(obj.get("validThrough")):
                continue

            title = clean_text(obj.get("title") or obj.get("name"))
            description = clean_text(obj.get("description"))
            if not is_target_job(title, description):
                continue

            location = parse_location(obj.get("jobLocation"))
            remote_requirements = parse_location(obj.get("applicantLocationRequirements"))
            job_location_type = clean_text(obj.get("jobLocationType"))
            if not location and remote_requirements:
                location = remote_requirements

            application_url = canonicalize_url(
                urljoin(page_url, clean_text(obj.get("url")) or page_url)
            )
            if not application_url:
                application_url = canonicalize_url(page_url)

            work_mode = detect_work_mode(
                job_location_type,
                location,
                title,
                description[:900],
            )
            if "telecommute" in job_location_type.casefold():
                work_mode = "Remote"

            jobs.append(
                Job(
                    company=company_name,
                    job_title=title,
                    location=location or "See posting",
                    work_mode=work_mode,
                    employment_type=clean_text(obj.get("employmentType")),
                    date_posted=clean_text(obj.get("datePosted")),
                    status="Active",
                    source=source,
                    application_url=application_url,
                    description=description,
                ).finalize()
            )

    return dedupe_jobs(jobs)


def extract_fallback_job(
    html: str,
    page_url: str,
    company_name: str,
    source: str,
) -> Job | None:
    soup = BeautifulSoup(html, "html.parser")
    page_text = clean_text(soup.get_text(" ", strip=True))
    if CLOSED_JOB_RE.search(page_text[:5000]):
        return None

    title_candidates = []
    for selector in [
        "h1",
        "[data-automation-id='jobPostingHeader']",
        "[data-automation-id='job-title']",
        ".job-title",
        "meta[property='og:title']",
        "title",
    ]:
        node = soup.select_one(selector)
        if not node:
            continue
        if node.name == "meta":
            title_candidates.append(clean_text(node.get("content")))
        else:
            title_candidates.append(clean_text(node.get_text(" ", strip=True)))

    title = next((item for item in title_candidates if is_target_job(item, page_text)), "")
    if not title:
        return None

    location = ""
    for selector in [
        "[data-automation-id='locations']",
        "[data-automation-id='location']",
        "[class*='job-location' i]",
        "[class*='location' i]",
        "[itemprop='jobLocation']",
    ]:
        node = soup.select_one(selector)
        if node:
            candidate = clean_text(node.get_text(" ", strip=True))
            if candidate and len(candidate) < 220:
                location = candidate
                break

    date_posted = ""
    date_node = soup.select_one("time[datetime], [itemprop='datePosted']")
    if date_node:
        date_posted = clean_text(date_node.get("datetime") or date_node.get_text(" ", strip=True))

    employment_type = ""
    match = re.search(
        r"\b(?:employment type|job type)\s*[:\-]?\s*"
        r"(full[- ]time|part[- ]time|contract|temporary|internship)",
        page_text,
        re.IGNORECASE,
    )
    if match:
        employment_type = match.group(1).title()

    canonical = soup.select_one("link[rel='canonical']")
    canonical_url = canonicalize_url(
        urljoin(page_url, canonical.get("href")) if canonical and canonical.get("href") else page_url
    )

    return Job(
        company=company_name,
        job_title=title,
        location=location or "See posting",
        work_mode=detect_work_mode(title, location, page_text[:1800]),
        employment_type=employment_type,
        date_posted=date_posted,
        status="Active",
        source=source,
        application_url=canonical_url,
        description=page_text[:3000],
    ).finalize()


def allowed_domains(company: dict[str, Any]) -> set[str]:
    values = company.get("allowed_domains") or []
    domains = {urlparse(company["careers_url"]).netloc.casefold()}
    domains.update(str(value).casefold() for value in values)
    return {domain.removeprefix("www.") for domain in domains if domain}


def domain_allowed(url: str, domains: set[str]) -> bool:
    domain = urlparse(url).netloc.casefold().removeprefix("www.")
    return any(domain == allowed or domain.endswith("." + allowed) for allowed in domains)


def extract_anchor_title(anchor: Any) -> str:
    values = [
        anchor.get("aria-label", ""),
        anchor.get("title", ""),
        anchor.get_text(" ", strip=True),
    ]
    parent = anchor.find_parent(["article", "li", "tr", "div"])
    if parent:
        heading = parent.find(["h1", "h2", "h3", "h4", "h5"])
        if heading:
            values.insert(0, heading.get_text(" ", strip=True))
    for value in values:
        text = clean_text(value)
        if text and len(text) <= 220:
            return text
    return ""


def looks_like_detail_url(url: str, detail_regex: str = "") -> bool:
    canonical = canonicalize_url(url)
    if not canonical:
        return False
    parsed = urlparse(canonical)
    combined = parsed.path + ("?" + parsed.query if parsed.query else "")
    if detail_regex:
        try:
            return bool(re.search(detail_regex, combined, re.IGNORECASE))
        except re.error:
            pass
    if LISTING_URL_RE.search(combined) and not JOB_URL_HINT_RE.search(combined):
        return False
    return bool(JOB_URL_HINT_RE.search(combined))


def slug_text(url: str) -> str:
    parsed = urlparse(url)
    value = parsed.path.rsplit("/", 1)[-1]
    return clean_text(re.sub(r"[-_]+", " ", value))


def extract_candidate_links_from_html(
    html: str,
    base_url: str,
    company: dict[str, Any],
) -> list[tuple[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    domains = allowed_domains(company)
    detail_regex = clean_text(company.get("detail_url_regex"))
    candidates: list[tuple[str, str, int]] = []

    for anchor in soup.select("a[href]"):
        raw = clean_text(anchor.get("href"))
        if not raw or raw.startswith(("mailto:", "tel:", "javascript:")):
            continue
        url = canonicalize_url(urljoin(base_url, raw))
        if not url or not domain_allowed(url, domains):
            continue
        title = extract_anchor_title(anchor)
        if not looks_like_detail_url(url, detail_regex):
            continue
        priority = 0
        if is_target_job(title):
            priority += 5
        if is_target_job(slug_text(url)):
            priority += 3
        if detail_regex:
            priority += 1
        candidates.append((url, title, priority))

    candidates.sort(key=lambda item: item[2], reverse=True)
    return dedupe_link_pairs([(url, title) for url, title, _ in candidates])[:MAX_CANDIDATES_PER_COMPANY]


def extract_candidate_links_from_pairs(
    pairs: list[tuple[str, str]],
    company: dict[str, Any],
) -> list[tuple[str, str]]:
    domains = allowed_domains(company)
    detail_regex = clean_text(company.get("detail_url_regex"))
    scored: list[tuple[str, str, int]] = []
    for raw_url, raw_title in pairs:
        url = canonicalize_url(raw_url)
        if not url or not domain_allowed(url, domains):
            continue
        if not looks_like_detail_url(url, detail_regex):
            continue
        title = clean_text(raw_title)
        priority = 5 if is_target_job(title) else 0
        if is_target_job(slug_text(url)):
            priority += 3
        scored.append((url, title, priority))
    scored.sort(key=lambda item: item[2], reverse=True)
    return dedupe_link_pairs([(url, title) for url, title, _ in scored])[:MAX_CANDIDATES_PER_COMPANY]


def discover_pagination_links(html: str, page_url: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    current_domain = urlparse(page_url).netloc.casefold()
    links: list[str] = []
    for anchor in soup.select("a[href]"):
        href = canonicalize_url(urljoin(page_url, clean_text(anchor.get("href"))))
        if not href or urlparse(href).netloc.casefold() != current_domain:
            continue
        text = clean_text(anchor.get_text(" ", strip=True))
        rel = " ".join(anchor.get("rel", []))
        if rel.casefold() == "next" or text.casefold() in {"next", "next page", ">", "›"}:
            links.append(href)
            continue
        parsed = urlparse(href)
        if re.search(r"/(?:page/)?\d+/?$", parsed.path) and parsed.query:
            links.append(href)
    return list(dict.fromkeys(links))[:MAX_LISTING_PAGES]


def dedupe_link_pairs(values: Iterable[tuple[str, str]]) -> list[tuple[str, str]]:
    output: list[tuple[str, str]] = []
    seen: set[str] = set()
    for url, title in values:
        key = canonicalize_url(url)
        if not key or key in seen:
            continue
        seen.add(key)
        output.append((key, clean_text(title)))
    return output


def fetch_text(session: requests.Session, url: str) -> str:
    response = session.get(url, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    content_type = response.headers.get("content-type", "")
    if "application/json" in content_type:
        return response.text
    response.encoding = response.encoding or response.apparent_encoding
    return response.text


def parse_detail_page(
    session: requests.Session,
    browser: BrowserCollector,
    url: str,
    company_name: str,
    source: str,
) -> list[Job]:
    html = ""
    try:
        html = fetch_text(session, url)
    except requests.RequestException:
        pass

    jobs = extract_jsonld_jobs(html, url, company_name, source) if html else []
    if jobs:
        return jobs

    fallback = extract_fallback_job(html, url, company_name, source) if html else None
    if fallback:
        return [fallback]

    rendered = browser.render_html(url) if browser.enabled else ""
    if rendered:
        jobs = extract_jsonld_jobs(rendered, url, company_name, source)
        if jobs:
            return jobs
        fallback = extract_fallback_job(rendered, url, company_name, source)
        if fallback:
            return [fallback]
    return []


def discover_sitemap_candidates(
    session: requests.Session,
    company: dict[str, Any],
) -> list[tuple[str, str]]:
    careers_url = company["careers_url"]
    parsed = urlparse(careers_url)
    origin = f"{parsed.scheme}://{parsed.netloc}"
    sitemap_urls = [urljoin(origin, "/sitemap.xml"), urljoin(origin, "/sitemap_index.xml")]

    try:
        robots = fetch_text(session, urljoin(origin, "/robots.txt"))
        for line in robots.splitlines():
            if line.casefold().startswith("sitemap:"):
                sitemap_urls.append(line.split(":", 1)[1].strip())
    except requests.RequestException:
        pass

    discovered_sitemaps: list[str] = []
    candidate_urls: list[tuple[str, str]] = []
    domains = allowed_domains(company)
    detail_regex = clean_text(company.get("detail_url_regex"))

    for sitemap_url in list(dict.fromkeys(sitemap_urls))[:8]:
        try:
            xml = fetch_text(session, sitemap_url)
        except requests.RequestException:
            continue
        soup = BeautifulSoup(xml, "xml")
        nested = [clean_text(loc.get_text()) for loc in soup.select("sitemap > loc")]
        discovered_sitemaps.extend(nested[:8])
        for loc in soup.select("url > loc"):
            url = canonicalize_url(clean_text(loc.get_text()))
            if not url or not domain_allowed(url, domains):
                continue
            if looks_like_detail_url(url, detail_regex) and (
                is_target_job(slug_text(url)) or detail_regex
            ):
                candidate_urls.append((url, slug_text(url)))
                if len(candidate_urls) >= MAX_SITEMAP_URLS:
                    return dedupe_link_pairs(candidate_urls)

    for sitemap_url in list(dict.fromkeys(discovered_sitemaps))[:8]:
        try:
            xml = fetch_text(session, sitemap_url)
        except requests.RequestException:
            continue
        soup = BeautifulSoup(xml, "xml")
        for loc in soup.select("url > loc"):
            url = canonicalize_url(clean_text(loc.get_text()))
            if not url or not domain_allowed(url, domains):
                continue
            if looks_like_detail_url(url, detail_regex) and (
                is_target_job(slug_text(url)) or detail_regex
            ):
                candidate_urls.append((url, slug_text(url)))
                if len(candidate_urls) >= MAX_SITEMAP_URLS:
                    return dedupe_link_pairs(candidate_urls)

    return dedupe_link_pairs(candidate_urls)


def collect_smartrecruiters(
    session: requests.Session,
    browser: BrowserCollector,
    company: dict[str, Any],
) -> tuple[list[Job], str]:
    careers_url = company["careers_url"]
    identifier = clean_text(company.get("company_identifier"))
    if not identifier:
        parsed = urlparse(careers_url)
        identifier = parsed.path.strip("/").split("/")[0]
    if not identifier:
        return collect_auto(session, browser, company)

    api_url = f"https://api.smartrecruiters.com/v1/companies/{identifier}/postings"
    jobs: list[Job] = []
    offset = 0
    try:
        while offset < 500:
            response = session.get(
                api_url,
                params={"limit": 100, "offset": offset, "destination": "PUBLIC"},
                timeout=REQUEST_TIMEOUT,
                headers={"Accept": "application/json"},
            )
            if response.status_code in {401, 403, 404}:
                break
            response.raise_for_status()
            payload = response.json()
            postings = payload.get("content") or payload.get("postings") or []
            if not postings:
                break

            for posting in postings:
                title = clean_text(posting.get("name") or posting.get("title"))
                if not is_target_job(title):
                    continue
                posting_id = clean_text(posting.get("id") or posting.get("uuid"))
                details: dict[str, Any] = posting
                if posting_id:
                    detail_response = session.get(
                        f"{api_url}/{posting_id}",
                        timeout=REQUEST_TIMEOUT,
                        headers={"Accept": "application/json"},
                    )
                    if detail_response.ok:
                        details = detail_response.json()

                location_value = details.get("location") or posting.get("location") or {}
                location = parse_location(location_value)
                description = clean_text(
                    details.get("jobAd", {}).get("sections", {})
                    if isinstance(details.get("jobAd"), dict)
                    else details.get("description")
                )
                apply_url = clean_text(
                    details.get("applyUrl")
                    or details.get("ref")
                    or posting.get("ref")
                )
                if not apply_url and posting_id:
                    apply_url = f"https://jobs.smartrecruiters.com/{identifier}/{posting_id}"

                jobs.append(
                    Job(
                        company=company["name"],
                        job_title=title,
                        location=location or "See posting",
                        work_mode=detect_work_mode(title, location, description),
                        employment_type=clean_text(
                            details.get("typeOfEmployment")
                            or posting.get("typeOfEmployment")
                        ),
                        date_posted=clean_text(
                            details.get("releasedDate") or posting.get("releasedDate")
                        ),
                        status="Active",
                        source="SmartRecruiters API",
                        application_url=apply_url,
                        description=description,
                    ).finalize()
                )

            if len(postings) < 100:
                break
            offset += 100

    except (requests.RequestException, ValueError, TypeError):
        jobs = []

    if jobs:
        return dedupe_jobs(jobs), "SmartRecruiters API"
    return collect_auto(session, browser, company)


def collect_auto(
    session: requests.Session,
    browser: BrowserCollector,
    company: dict[str, Any],
) -> tuple[list[Job], str]:
    listing_urls = company.get("search_urls") or [company["careers_url"]]
    listing_queue = [canonicalize_url(str(url)) for url in listing_urls if canonicalize_url(str(url))]
    visited_listings: set[str] = set()
    candidate_pairs: list[tuple[str, str]] = []
    method_parts: list[str] = []

    while listing_queue and len(visited_listings) < MAX_LISTING_PAGES:
        listing_url = listing_queue.pop(0)
        if listing_url in visited_listings:
            continue
        visited_listings.add(listing_url)

        html = ""
        try:
            html = fetch_text(session, listing_url)
            method_parts.append("HTTP")
        except requests.RequestException:
            pass

        if html:
            candidate_pairs.extend(extract_candidate_links_from_html(html, listing_url, company))
            for pagination_url in discover_pagination_links(html, listing_url):
                if pagination_url not in visited_listings:
                    listing_queue.append(pagination_url)

    candidate_pairs = dedupe_link_pairs(candidate_pairs)

    if not candidate_pairs and browser.enabled:
        configured_terms = company.get("search_terms")
        if isinstance(configured_terms, list) and configured_terms:
            search_terms = [clean_text(term) for term in configured_terms if clean_text(term)]
        else:
            search_terms = ["analyst", "data", "AI"]

        for listing_url in listing_urls[:3]:
            terms = search_terms if company.get("browser_search", True) else [None]
            for search_term in terms:
                pairs, rendered_html = browser.collect_links(
                    str(listing_url),
                    search_term=search_term,
                )
                method_parts.append("Playwright")
                candidate_pairs.extend(extract_candidate_links_from_pairs(pairs, company))
                if rendered_html:
                    candidate_pairs.extend(
                        extract_candidate_links_from_html(
                            rendered_html,
                            str(listing_url),
                            company,
                        )
                    )
                candidate_pairs = dedupe_link_pairs(candidate_pairs)
                if len(candidate_pairs) >= MAX_CANDIDATES_PER_COMPANY:
                    break
            if len(candidate_pairs) >= MAX_CANDIDATES_PER_COMPANY:
                break

    candidate_pairs = dedupe_link_pairs(candidate_pairs)

    if len(candidate_pairs) < 2:
        sitemap_pairs = discover_sitemap_candidates(session, company)
        if sitemap_pairs:
            method_parts.append("sitemap")
            candidate_pairs.extend(sitemap_pairs)

    candidate_pairs = dedupe_link_pairs(candidate_pairs)[:MAX_CANDIDATES_PER_COMPANY]
    jobs: list[Job] = []
    source = "+".join(dict.fromkeys(method_parts)) or "auto"

    for url, anchor_title in candidate_pairs:
        if anchor_title and not is_target_job(anchor_title) and not is_target_job(slug_text(url)):
            # We still validate some generic links because many sites label links "View job".
            if anchor_title.casefold() not in {"view job", "learn more", "apply", "details", ""}:
                continue
        jobs.extend(parse_detail_page(session, browser, url, company["name"], source))
        time.sleep(0.08)

    return dedupe_jobs(jobs), source


def dedupe_jobs(jobs: Iterable[Job]) -> list[Job]:
    output: list[Job] = []
    seen: set[str] = set()
    for job in jobs:
        job.finalize()
        key = canonicalize_url(job.application_url) or job.job_id
        if not key or key in seen:
            continue
        seen.add(key)
        output.append(job)
    return output


def load_companies() -> list[dict[str, Any]]:
    with CONFIG_FILE.open("r", encoding="utf-8") as file:
        companies = json.load(file)
    if not isinstance(companies, list):
        raise ValueError("config/companies.json must contain a JSON list.")

    validated: list[dict[str, Any]] = []
    for index, company in enumerate(companies, start=1):
        if not isinstance(company, dict):
            raise ValueError(f"Company #{index} is not a JSON object.")
        if not clean_text(company.get("name")):
            raise ValueError(f"Company #{index} is missing 'name'.")
        if not clean_text(company.get("careers_url")):
            raise ValueError(f"{company.get('name')} is missing 'careers_url'.")
        if not canonicalize_url(str(company["careers_url"])):
            raise ValueError(f"{company.get('name')} has an invalid careers_url.")
        company.setdefault("collector", "auto")
        company.setdefault("enabled", True)
        validated.append(company)
    return validated


def load_previous_state() -> dict[str, dict[str, Any]]:
    if not STATE_JSON.exists():
        return {}
    try:
        with STATE_JSON.open("r", encoding="utf-8") as file:
            value = json.load(file)
        return value if isinstance(value, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def apply_state(jobs: list[Job], previous: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    current_state: dict[str, dict[str, Any]] = {}

    for job in jobs:
        old = previous.get(job.job_id, {})
        job.first_seen = clean_text(old.get("first_seen")) or now
        job.last_seen = now
        job.is_new = job.job_id not in previous
        current_state[job.job_id] = {
            "first_seen": job.first_seen,
            "last_seen": job.last_seen,
            "company": job.company,
            "job_title": job.job_title,
            "application_url": job.application_url,
        }
    return current_state


def make_excel_clickable(path: Path) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active
    headers = {cell.value: cell.column for cell in worksheet[1]}
    url_column = headers.get("application_url")
    if url_column:
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=url_column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = {
        "A": 24,
        "B": 38,
        "C": 24,
        "D": 14,
        "E": 16,
        "F": 14,
        "G": 12,
        "H": 20,
        "I": 70,
        "J": 14,
        "K": 22,
        "L": 22,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    workbook.save(path)
    workbook.close()


def save_outputs(jobs: list[Job], health: list[CompanyHealth], run_started: datetime) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    DASHBOARD_DATA_DIR.mkdir(parents=True, exist_ok=True)

    rows = [asdict(job) for job in jobs]
    rows.sort(key=lambda item: (not item["is_new"], item["company"], item["job_title"]))

    with CURRENT_JSON.open("w", encoding="utf-8") as file:
        json.dump(rows, file, ensure_ascii=False, indent=2)
    shutil.copy2(CURRENT_JSON, DASHBOARD_JOBS_JSON)

    dataframe = pd.DataFrame(rows)
    ordered_columns = [
        "company",
        "job_title",
        "location",
        "work_mode",
        "employment_type",
        "date_posted",
        "status",
        "source",
        "application_url",
        "is_new",
        "first_seen",
        "last_seen",
        "job_id",
        "description",
    ]
    dataframe = dataframe.reindex(columns=ordered_columns)
    dataframe.to_csv(CURRENT_CSV, index=False)
    dataframe.to_excel(EXCEL_FILE, index=False)
    make_excel_clickable(EXCEL_FILE)

    health_rows = [asdict(item) for item in health]
    with HEALTH_JSON.open("w", encoding="utf-8") as file:
        json.dump(health_rows, file, ensure_ascii=False, indent=2)

    now = datetime.now(timezone.utc)
    companies_with_jobs = sum(1 for item in health if item.jobs_found > 0)
    meta = {
        "last_updated": now.isoformat(timespec="seconds"),
        "run_duration_seconds": round((now - run_started).total_seconds(), 2),
        "total_jobs": len(rows),
        "new_jobs": sum(1 for item in rows if item["is_new"]),
        "companies_configured": len(health),
        "companies_with_jobs": companies_with_jobs,
        "companies_zero_results": sum(1 for item in health if item.status == "zero"),
        "companies_failed": sum(1 for item in health if item.status == "error"),
        "health": health_rows,
    }
    with DASHBOARD_META_JSON.open("w", encoding="utf-8") as file:
        json.dump(meta, file, ensure_ascii=False, indent=2)


def collect_company(
    session: requests.Session,
    browser: BrowserCollector,
    company: dict[str, Any],
) -> tuple[list[Job], str]:
    collector = clean_text(company.get("collector")).casefold() or "auto"
    host = urlparse(company["careers_url"]).netloc.casefold()
    if collector == "smartrecruiters" or "smartrecruiters.com" in host:
        return collect_smartrecruiters(session, browser, company)
    return collect_auto(session, browser, company)


def run(no_browser: bool = False, company_filter: str = "") -> int:
    companies = load_companies()
    enabled_companies = [
        company
        for company in companies
        if company.get("enabled", True)
    ]

    if company_filter:
        requested_company = company_filter.strip().casefold()
        enabled_companies = [
            company
            for company in enabled_companies
            if clean_text(company.get("name")).casefold() == requested_company
        ]

        if not enabled_companies:
            print(f"Company not found: {company_filter}")
            print("Use one of these exact company names:")
            for company in companies:
                if company.get("enabled", True):
                    print(f"- {clean_text(company.get('name'))}")
            return 1

        print(f"Running only: {enabled_companies[0]['name']}")

    print(f"Companies selected: {len(enabled_companies)}")
    print()

    previous_state = load_previous_state()
    all_jobs: list[Job] = []
    health: list[CompanyHealth] = []
    run_started = datetime.now(timezone.utc)
    session = create_session()

    with BrowserCollector(enabled=not no_browser) as browser:
        for index, company in enumerate(enabled_companies, start=1):
            name = company["name"]
            started = time.perf_counter()
            print(f"[{index}/{len(enabled_companies)}] Checking {name}...")
            try:
                company_jobs, method = collect_company(session, browser, company)
                company_jobs = dedupe_jobs(company_jobs)
                all_jobs.extend(company_jobs)
                elapsed = round(time.perf_counter() - started, 2)
                status = "ok" if company_jobs else "zero"
                message = "" if company_jobs else "No verified Data/AI job detail pages were found."
                health.append(
                    CompanyHealth(
                        company=name,
                        status=status,
                        jobs_found=len(company_jobs),
                        method=method,
                        duration_seconds=elapsed,
                        message=message,
                        careers_url=company["careers_url"],
                    )
                )
                print(f"Found {len(company_jobs)} verified Data/AI jobs via {method}")
            except Exception as error:  # One company must never stop the complete run.
                elapsed = round(time.perf_counter() - started, 2)
                health.append(
                    CompanyHealth(
                        company=name,
                        status="error",
                        jobs_found=0,
                        method=clean_text(company.get("collector")) or "auto",
                        duration_seconds=elapsed,
                        message=f"{type(error).__name__}: {error}",
                        careers_url=company["careers_url"],
                    )
                )
                print(f"Could not collect {name}: {type(error).__name__}: {error}")
            print()

    all_jobs = dedupe_jobs(all_jobs)
    current_state = apply_state(all_jobs, previous_state)
    save_outputs(all_jobs, health, run_started)

    with STATE_JSON.open("w", encoding="utf-8") as file:
        json.dump(current_state, file, ensure_ascii=False, indent=2)

    print("----------------------------------")
    print("Job collection finished")
    print("----------------------------------")
    print(f"Configured companies: {len(enabled_companies)}")
    print(f"Companies returning jobs: {sum(1 for item in health if item.jobs_found > 0)}")
    print(f"Companies returning 0 jobs: {sum(1 for item in health if item.status == 'zero')}")
    print(f"Companies failed: {sum(1 for item in health if item.status == 'error')}")
    print(f"Unique jobs saved: {len(all_jobs)}")
    print(f"New jobs this run: {sum(1 for job in all_jobs if job.is_new)}")
    print()
    print(f"JSON created: {CURRENT_JSON.relative_to(ROOT)}")
    print(f"CSV created: {CURRENT_CSV.relative_to(ROOT)}")
    print(f"Excel created: {EXCEL_FILE.relative_to(ROOT)}")
    print(f"Dashboard data: {DASHBOARD_DATA_DIR.relative_to(ROOT)}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Collect public Data/AI job postings.")
    parser.add_argument(
        "--no-browser",
        action="store_true",
        help="Use requests only and skip the Playwright browser fallback.",
    )
    parser.add_argument(
        "--company",
        default="",
        help="Run one company using its exact name from companies.json.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    try:
        raise SystemExit(run(no_browser=arguments.no_browser, company_filter=arguments.company))
    except KeyboardInterrupt:
        print("Stopped by user.")
        raise SystemExit(130)
    except Exception as exc:
        print(f"Fatal error: {type(exc).__name__}: {exc}", file=sys.stderr)
        raise SystemExit(1)
